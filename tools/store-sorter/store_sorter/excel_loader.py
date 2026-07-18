"""Excel マスターファイルの読み込み（openpyxl）。"""

from __future__ import annotations

from pathlib import Path
from typing import List, Optional, Union

from openpyxl import load_workbook

PathLike = Union[str, Path]


def sheet_names(path: PathLike) -> List[str]:
    """ブック内のシート名一覧を返す。"""
    wb = load_workbook(path, read_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def read_headers(path: PathLike, sheet_name: Optional[str] = None) -> List[str]:
    """1 行目（ヘッダー行）を文字列リストとして返す。

    空セルは空文字列に、前後の空白は除去する。
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            return [("" if c is None else str(c)).strip() for c in row]
        return []
    finally:
        wb.close()


def read_column(
    path: PathLike, col_index: int, sheet_name: Optional[str] = None
) -> List[str]:
    """指定列（0 始まり）の 2 行目以降の値を返す。

    空セルを除外し、重複を（出現順を維持して）取り除いた店舗名マスターを返す。
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        seen = set()
        result: List[str] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if col_index >= len(row):
                continue
            value = row[col_index]
            if value is None:
                continue
            text = str(value).strip()
            if not text or text in seen:
                continue
            seen.add(text)
            result.append(text)
        return result
    finally:
        wb.close()
